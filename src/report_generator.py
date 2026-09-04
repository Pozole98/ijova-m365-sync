"""
Generador exhaustivo de reportes en consola, Markdown y archivos CSV segmentados.
Exporta existentes.csv, nuevos.csv, conflictos.csv, invalidos.csv y discrepancias.csv.
"""
import os
import csv
from datetime import datetime, timezone
from typing import List, Dict, Any, Optional
from tabulate import tabulate
from src.models import StudentRecord, ClassificationEnum, DomainStatus


def generate_reports(
    students: List[StudentRecord],
    discrepancies: List[Dict[str, str]],
    domain_status: Optional[DomainStatus] = None,
    admin_upn: Optional[str] = None,
    output_dir: str = "reports",
    is_dry_run: bool = True
) -> Dict[str, str]:
    """
    Genera y guarda los reportes estructurados, archivos CSV segmentados y el resumen de auditoría Markdown.
    Incluye rastro de auditoría: Timestamp UTC y Admin UPN.
    """
    os.makedirs(output_dir, exist_ok=True)
    generated_files: Dict[str, str] = {}
    timestamp_utc = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")

    # Separate students by classification
    existentes = [s for s in students if s.classification == ClassificationEnum.EXISTENTE]
    nuevos = [s for s in students if s.classification == ClassificationEnum.NUEVO]
    conflictos = [s for s in students if s.classification == ClassificationEnum.CONFLICTO]
    invalidos = [s for s in students if s.classification == ClassificationEnum.INVALIDO]

    # Metrics
    total_registros = len(students)
    matriculas_unicas = len(set(s.matricula for s in students if s.matricula))
    
    # Personas potencialmente únicas
    import unicodedata
    def norm_name(s: StudentRecord) -> str:
        t = unicodedata.normalize('NFKD', f"{s.nombres} {s.apellido_paterno} {s.apellido_materno}").encode('ASCII', 'ignore').decode('utf-8')
        return " ".join(t.upper().split())
    personas_unicas = len(set(norm_name(s) for s in students))

    # Helper function to write CSV
    def write_student_csv(filename: str, records: List[StudentRecord], extra_cols: bool = False):
        filepath = os.path.join(output_dir, filename)
        fieldnames = [
            "row_index", "matricula", "nombres", "apellido_paterno", "apellido_materno",
            "display_name", "nivel", "grado_semestre", "estatus",
            "upn_normalized", "alias_normalized", "mail_nickname", "classification"
        ]
        if extra_cols:
            fieldnames.extend(["motivo_bloqueo_o_error", "entra_id_match"])

        with open(filepath, "w", encoding="utf-8-sig", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            for s in records:
                row_dict = {
                    "row_index": s.row_index,
                    "matricula": s.matricula,
                    "nombres": s.nombres,
                    "apellido_paterno": s.apellido_paterno,
                    "apellido_materno": s.apellido_materno,
                    "display_name": s.display_name,
                    "nivel": s.nivel,
                    "grado_semestre": s.grado_semestre,
                    "estatus": s.estatus,
                    "upn_normalized": s.upn_normalized,
                    "alias_normalized": s.alias_normalized,
                    "mail_nickname": s.mail_nickname,
                    "classification": s.classification.value
                }
                if extra_cols:
                    issues_str = " | ".join([f"[{i.code}] {i.message}" for i in s.issues])
                    row_dict["motivo_bloqueo_o_error"] = issues_str
                    row_dict["entra_id_match"] = s.entra_id_match or ""
                writer.writerow(row_dict)

        generated_files[filename] = filepath

    # 1. Export CSVs
    write_student_csv("existentes.csv", existentes, extra_cols=True)
    write_student_csv("nuevos.csv", nuevos, extra_cols=True)
    write_student_csv("conflictos.csv", conflictos, extra_cols=True)
    write_student_csv("invalidos.csv", invalidos, extra_cols=True)

    # 2. Export Discrepancias CSV
    disc_path = os.path.join(output_dir, "discrepancias.csv")
    with open(disc_path, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=["matricula", "upn", "campo", "valor_entra", "valor_excel"])
        writer.writeheader()
        for d in discrepancies:
            writer.writerow(d)
    generated_files["discrepancias.csv"] = disc_path

    # 3. Export Markdown Summary for Audit Evidence
    md_summary_path = os.path.join(output_dir, "dry_run_audit_summary.md")
    with open(md_summary_path, "w", encoding="utf-8") as f:
        f.write("# Rastro de Auditoría y Resumen de Simulación (Dry-Run)\n\n")
        f.write(f"- **Fecha y Hora de Ejecución:** `{timestamp_utc}`\n")
        f.write(f"- **Administrador Responsable (UPN):** `{admin_upn or 'No autenticado / Validación local'}`\n")
        f.write(f"- **Modo de Operación:** `{'DRY-RUN (Simulación 100% solo lectura)' if is_dry_run else 'VALIDACIÓN LOCAL'}`\n")
        f.write(f"- **Dominio Validado:** `{domain_status.domain_name if domain_status else 'N/A'}` (Verificado: `{domain_status.is_verified if domain_status else 'N/A'}`)\n\n")
        f.write("## 📊 Métricas de Sincronización\n\n")
        f.write(f"| Métrica | Cantidad |\n| :--- | :---: |\n")
        f.write(f"| Registros Administrativos Totales | **{total_registros}** |\n")
        f.write(f"| Matrículas Únicas en Hoja | **{matriculas_unicas}** |\n")
        f.write(f"| Personas Potencialmente Únicas | **{personas_unicas}** |\n")
        f.write(f"| Cuentas EXISTENTES en Entra ID (Intactas) | **{len(existentes)}** |\n")
        f.write(f"| Cuentas NUEVAS (Candidatas a creación) | **{len(nuevos)}** |\n")
        f.write(f"| CONFLICTOS (Bloqueados para creación) | **{len(conflictos)}** |\n")
        f.write(f"| INVÁLIDOS (Errores estructurales) | **{len(invalidos)}** |\n")
        f.write(f"| Discrepancias Detectadas | **{len(discrepancies)}** |\n\n")

        if conflictos:
            f.write("## ⚠️ Registros Bloqueados por Conflicto\n\n")
            f.write("| Fila | Matrícula | Nombre | UPN | Motivo de Conflicto |\n| :---: | :---: | :--- | :--- | :--- |\n")
            for c in conflictos:
                issues_msg = " ; ".join([f"[{i.code}] {i.message}" for i in c.issues if i.severity == "ERROR"])
                f.write(f"| {c.row_index} | {c.matricula} | {c.display_name} | {c.upn_normalized} | {issues_msg} |\n")
    generated_files["dry_run_audit_summary.md"] = md_summary_path

    # 4. Export Styled Excel (.xlsx) Report with Color Coding
    xlsx_path = os.path.join(output_dir, "reporte_sincronizacion_ijova.xlsx")
    try:
        export_styled_excel_report(
            students=students,
            discrepancies=discrepancies,
            output_path=xlsx_path,
            timestamp_utc=timestamp_utc,
            admin_upn=admin_upn,
            is_dry_run=is_dry_run
        )
        generated_files["reporte_sincronizacion_ijova.xlsx"] = xlsx_path
    except Exception as e:
        print(f"⚠️ Aviso: No se pudo generar el reporte Excel con estilos: {e}")

    # 5. Print Console Summary Table
    print("\n" + "=" * 80)
    mode_title = "SIMULACIÓN DRY RUN - MICROSOFT ENTRA ID" if is_dry_run else "VALIDACIÓN LOCAL DE EXCEL"
    print(f"📊 REPORTE DE RESULTADOS: {mode_title}")
    print("=" * 80)
    print(f"🕒 Timestamp de Ejecución: {timestamp_utc}")
    print(f"👤 Administrador Autenticado: {admin_upn or 'Validación Local / Sin Autenticación'}")
    print("-" * 80)

    summary_table = [
        ["Registros Administrativos Totales", total_registros],
        ["Matrículas Únicas en Hoja", matriculas_unicas],
        ["Personas Potencialmente Únicas", personas_unicas],
        ["----------------------------------------", "------"],
        ["Cuentas EXISTENTES en Entra ID (Omitidas sin cambios)", len(existentes)],
        ["Cuentas NUEVAS (Candidatas a creación futura)", len(nuevos)],
        ["CONFLICTOS (Bloqueados - Requieren decisión)", len(conflictos)],
        ["INVÁLIDOS (Errores de formato en hoja)", len(invalidos)],
        ["----------------------------------------", "------"],
        ["Discrepancias Hoja vs Entra ID", len(discrepancies)]
    ]
    print(tabulate(summary_table, headers=["Métrica", "Cantidad"], tablefmt="fancy_grid"))

    if domain_status:
        print("\n🌐 Estado del Dominio Institucional:")
        print(f"   - Dominio: {domain_status.domain_name}")
        print(f"   - Verificado: {'✅ SÍ' if domain_status.is_verified else '❌ NO (BLOQUEO)'}")
        print(f"   - Tipo de Autenticación: {domain_status.authentication_type}")
        if domain_status.is_blocked:
            print(f"   - \033[1;31mALERTA:\033[0m {domain_status.block_reason}")

    if conflictos:
        print("\n⚠️ REGISTROS EN CONFLICTO (BLOQUEADOS PARA CREACIÓN):")
        conflict_rows = []
        for c in conflictos:
            issues_msg = " ; ".join([f"[{i.code}] {i.message}" for i in c.issues if i.severity == "ERROR"])
            conflict_rows.append([c.row_index, c.matricula, c.display_name, c.upn_normalized, issues_msg])
        print(tabulate(conflict_rows, headers=["Fila", "Matrícula", "Nombre", "UPN", "Motivo de Conflicto"], tablefmt="grid"))

    print("\n📁 Archivos de reporte generados en:", os.path.abspath(output_dir))
    for name, path in generated_files.items():
        print(f"   • {name}: {path}")

    # 6. Registrar en auditoría
    try:
        from src.audit_logger import log_audit_event
        log_audit_event(
            action="REPORT_GENERATE",
            target=f"{total_registros} alumnos",
            admin=admin_upn or "Local",
            status="SUCCESS",
            details=f"Existentes: {len(existentes)} | Nuevos: {len(nuevos)} | Conflictos: {len(conflictos)} | Inválidos: {len(invalidos)}"
        )
    except Exception:
        pass

    return generated_files


def export_styled_excel_report(
    students: List[StudentRecord],
    discrepancies: List[Dict[str, str]],
    output_path: str,
    timestamp_utc: str,
    admin_upn: Optional[str] = None,
    is_dry_run: bool = True
) -> str:
    """
    Genera un libro Excel (.xlsx) estructurado y con diseño institucional profesional:
    - Pestaña 'Resumen Ejecutivo' con métricas globales y branding del IJOVA.
    - Pestañas coloreadas por estatus (Existentes: Verde, Nuevos: Azul, Conflictos: Naranja, Inválidos: Rojo).
    - Auto-ajuste de anchos de columna para legibilidad inmediata.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    thin_border = Border(
        left=Side(style='thin', color='D3D3D3'),
        right=Side(style='thin', color='D3D3D3'),
        top=Side(style='thin', color='D3D3D3'),
        bottom=Side(style='thin', color='D3D3D3')
    )

    # 1. Sheet Resumen Ejecutivo
    ws_resumen = wb.create_sheet(title="Resumen Ejecutivo")
    ws_resumen.views.sheetView[0].showGridLines = True

    # Title
    ws_resumen.merge_cells("A1:D1")
    ws_resumen["A1"] = "INSTITUTO DE DESARROLLO INTEGRAL LIC. JOSÉ VASCONCELOS (IJOVA)"
    ws_resumen["A1"].font = Font(name="Arial", size=13, bold=True, color="FFFFFF")
    ws_resumen["A1"].fill = PatternFill(start_color="203A63", end_color="203A63", fill_type="solid")
    ws_resumen["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws_resumen.row_dimensions[1].height = 28

    ws_resumen.merge_cells("A2:D2")
    ws_resumen["A2"] = "INFORME EJECUTIVO DE SINCRONIZACIÓN Y AUDITORÍA - MICROSOFT 365"
    ws_resumen["A2"].font = Font(name="Arial", size=10.5, bold=True, color="FFFFFF")
    ws_resumen["A2"].fill = PatternFill(start_color="2E5288", end_color="2E5288", fill_type="solid")
    ws_resumen["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws_resumen.row_dimensions[2].height = 22

    # Metadata
    meta_rows = [
        ("Fecha y Hora UTC:", timestamp_utc),
        ("Administrador en Turno:", admin_upn or "Validación Local"),
        ("Modo de Operación:", "DRY-RUN (Simulación)" if is_dry_run else "PRODUCCIÓN (Aplicado)"),
        ("Total Registros Procesados:", len(students))
    ]
    for r_idx, (label, val) in enumerate(meta_rows, start=4):
        ws_resumen[f"A{r_idx}"] = label
        ws_resumen[f"A{r_idx}"].font = Font(name="Arial", size=10, bold=True, color="333333")
        ws_resumen[f"B{r_idx}"] = val
        ws_resumen[f"B{r_idx}"].font = Font(name="Arial", size=10)

    # Summary metrics table
    existentes = [s for s in students if s.classification == ClassificationEnum.EXISTENTE]
    nuevos = [s for s in students if s.classification == ClassificationEnum.NUEVO]
    conflictos = [s for s in students if s.classification == ClassificationEnum.CONFLICTO]
    invalidos = [s for s in students if s.classification == ClassificationEnum.INVALIDO]

    ws_resumen["A9"] = "Métrica de Clasificación"
    ws_resumen["B9"] = "Cantidad"
    ws_resumen["C9"] = "% del Total"
    ws_resumen["D9"] = "Estatus Operativo"

    for col in ["A9", "B9", "C9", "D9"]:
        ws_resumen[col].font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        ws_resumen[col].fill = PatternFill(start_color="203A63", end_color="203A63", fill_type="solid")
        ws_resumen[col].alignment = Alignment(horizontal="center", vertical="center")

    total_count = len(students) or 1
    metric_rows = [
        ("Alumnos Existentes (Sincronizados)", len(existentes), f"{(len(existentes)/total_count*100):.1f}%", "Activos en M365", "E8F5E9", "2E7D32"),
        ("Alumnos Nuevos (Por Aprovisionar)", len(nuevos), f"{(len(nuevos)/total_count*100):.1f}%", "Pendientes de Alta", "E3F2FD", "1565C0"),
        ("Conflictos (Bloqueados por Reglas)", len(conflictos), f"{(len(conflictos)/total_count*100):.1f}%", "Requiere Decisión", "FFF3E0", "E65100"),
        ("Registros Inválidos (Formato Erróneo)", len(invalidos), f"{(len(invalidos)/total_count*100):.1f}%", "Bloqueado", "FFEBEE", "C62828"),
        ("Discrepancias Detectadas", len(discrepancies), "-", "Aviso Informativo", "F3E5F5", "6A1B9A")
    ]

    for r_idx, (m_name, count, pct, st_desc, bg_color, text_color) in enumerate(metric_rows, start=10):
        ws_resumen[f"A{r_idx}"] = m_name
        ws_resumen[f"B{r_idx}"] = count
        ws_resumen[f"C{r_idx}"] = pct
        ws_resumen[f"D{r_idx}"] = st_desc

        ws_resumen[f"A{r_idx}"].font = Font(name="Arial", size=9.5, bold=True, color=text_color)
        ws_resumen[f"A{r_idx}"].fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
        ws_resumen[f"B{r_idx}"].alignment = Alignment(horizontal="center")
        ws_resumen[f"C{r_idx}"].alignment = Alignment(horizontal="center")
        ws_resumen[f"D{r_idx}"].alignment = Alignment(horizontal="center")

        for col_letter in ["A", "B", "C", "D"]:
            ws_resumen[f"{col_letter}{r_idx}"].border = thin_border

    def create_student_sheet(title, recs, header_color, row_bg_color, include_issues=False):
        ws = wb.create_sheet(title=title)
        ws.views.sheetView[0].showGridLines = True

        headers = ["Fila", "Matrícula", "Nombre Completo", "Nivel", "Grado / Semestre", "Correo Institucional (UPN)", "Alias", "Estatus Escolar"]
        if include_issues:
            headers.append("Motivo de Bloqueo / Observaciones")

        ws.append(headers)
        ws.row_dimensions[1].height = 24
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for r_idx, s in enumerate(recs, start=2):
            row_data = [
                s.row_index,
                s.matricula,
                s.display_name or f"{s.nombres} {s.apellido_paterno} {s.apellido_materno}".strip(),
                s.nivel,
                s.grado_semestre,
                s.upn_normalized or s.upn_raw,
                s.alias_normalized or s.alias_raw,
                s.estatus
            ]
            if include_issues:
                issues_text = " | ".join([f"[{i.code}] {i.message}" for i in s.issues])
                row_data.append(issues_text)

            ws.append(row_data)
            row_fill = PatternFill(start_color=row_bg_color, end_color=row_bg_color, fill_type="solid") if r_idx % 2 == 0 else None
            for col_idx in range(1, len(row_data) + 1):
                c = ws.cell(row=r_idx, column=col_idx)
                c.font = Font(name="Arial", size=9.5)
                c.border = thin_border
                if row_fill:
                    c.fill = row_fill
                if col_idx in [1, 2, 8]:
                    c.alignment = Alignment(horizontal="center")

        # Auto fit column widths
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = min(max(max_len + 3, 10), 45)

    create_student_sheet("Existentes Sincronizados", existentes, "2E7D32", "F1F8E9")
    create_student_sheet("Nuevos a Crear", nuevos, "1565C0", "E3F2FD")
    create_student_sheet("Conflictos Bloqueados", conflictos, "E65100", "FFF3E0", include_issues=True)
    create_student_sheet("Registros Invalidos", invalidos, "C62828", "FFEBEE", include_issues=True)

    if discrepancies:
        ws_disc = wb.create_sheet(title="Discrepancias")
        ws_disc.views.sheetView[0].showGridLines = True
        d_headers = ["Matrícula", "Correo Institucional", "Campo Afectado", "Valor en Entra ID", "Valor en Archivo Escolar"]
        ws_disc.append(d_headers)
        ws_disc.row_dimensions[1].height = 24
        for col_idx in range(1, len(d_headers) + 1):
            cell = ws_disc.cell(row=1, column=col_idx)
            cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="6A1B9A", end_color="6A1B9A", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for r_idx, d in enumerate(discrepancies, start=2):
            ws_disc.append([d.get("matricula", ""), d.get("upn", ""), d.get("campo", ""), d.get("valor_entra", ""), d.get("valor_excel", "")])
            for col_idx in range(1, 6):
                c = ws_disc.cell(row=r_idx, column=col_idx)
                c.font = Font(name="Arial", size=9.5)
                c.border = thin_border

        for col in ws_disc.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws_disc.column_dimensions[get_column_letter(col[0].column)].width = min(max(max_len + 3, 12), 40)

    for col in ws_resumen.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws_resumen.column_dimensions[get_column_letter(col[0].column)].width = min(max(max_len + 3, 14), 50)

    wb.save(output_path)
    return output_path
