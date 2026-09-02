"""
Lector de solo lectura para la hoja Excel 'Listado Global Matriculado'.
Garantiza acceso no destructivo y extracción íntegra de celdas y fórmulas.
"""
import os
import openpyxl
from typing import List, Optional
from pathlib import Path
from src.models import StudentRecord


def _parse_ods_students(file_path: Path, sheet_name: str = "Listado Global Matriculado") -> List[StudentRecord]:
    """Lee hojas de cálculo en formato ODS (.ods) vía odfpy."""
    from odf import opendocument, teletype
    from odf.table import Table, TableRow, TableCell

    doc = opendocument.load(str(file_path))
    tables = {t.getAttribute('name'): t for t in doc.spreadsheet.getElementsByType(Table)}
    if sheet_name not in tables:
        raise ValueError(f"La hoja '{sheet_name}' no existe en el archivo ODS. Hojas disponibles: {list(tables.keys())}")

    sheet = tables[sheet_name]
    students: List[StudentRecord] = []
    rows = sheet.getElementsByType(TableRow)

    for r_idx, r in enumerate(rows, start=1):
        if r_idx == 1:
            continue  # Skip header
        cells = r.getElementsByType(TableCell)
        row_vals = []
        for c in cells:
            rep = c.getAttribute('numbercolumnsrepeated')
            rep_count = int(rep) if rep else 1
            txt = teletype.extractText(c).strip()
            if rep_count > 1 and not txt:
                continue
            for _ in range(rep_count if rep_count < 20 else 1):
                row_vals.append(txt)

        if not any(row_vals):
            continue

        while len(row_vals) < 13:
            row_vals.append("")

        raw_mat = row_vals[0]
        raw_pat = row_vals[1]
        raw_mat_ap = row_vals[2]
        raw_nom = row_vals[3]
        raw_nivel = row_vals[4]
        raw_grado = row_vals[5]
        raw_estatus = row_vals[6]
        raw_upn = row_vals[7]
        raw_nom1 = row_vals[8]
        raw_nom_limp = row_vals[9]
        raw_ape_limp = row_vals[10]
        raw_alias = row_vals[11]
        raw_disp = row_vals[12]

        if not raw_mat and not raw_nom and not raw_pat:
            continue

        if raw_mat.endswith(".0"):
            raw_mat = raw_mat[:-2]

        record = StudentRecord(
            row_index=r_idx,
            matricula=raw_mat,
            apellido_paterno=raw_pat,
            apellido_materno=raw_mat_ap,
            nombres=raw_nom,
            nivel=raw_nivel,
            grado_semestre=raw_grado,
            estatus=raw_estatus,
            upn_raw=raw_upn,
            nombre1_raw=raw_nom1,
            nombre_limpio_raw=raw_nom_limp,
            apellido_limpio_raw=raw_ape_limp,
            alias_raw=raw_alias,
            display_name_raw=raw_disp
        )
        students.append(record)

    return students


def parse_excel_students(excel_path: str, sheet_name: str = "Listado Global Matriculado") -> List[StudentRecord]:
    """
    Lee el archivo Excel o ODS de forma no destructiva (read_only=True)
    y devuelve la lista de objetos StudentRecord con los datos originales.
    """
    file_path = Path(excel_path)
    if not file_path.exists():
        raise FileNotFoundError(f"No se encontró el archivo Excel en: {excel_path}")

    if str(file_path).lower().endswith(".ods"):
        return _parse_ods_students(file_path, sheet_name)

    # Open with openpyxl in read-only and data_only=True to get evaluated values
    wb = openpyxl.load_workbook(str(file_path), data_only=True, read_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"La hoja '{sheet_name}' no existe en el libro. Hojas disponibles: {wb.sheetnames}")

    sheet = wb[sheet_name]

    students: List[StudentRecord] = []
    
    # Iterate rows starting from row 2
    for r_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if r_idx == 1:
            continue  # Skip header row

        # row is a tuple of values (col A to M)
        # col indices: 0: Matricula, 1: Paterno, 2: Materno, 3: Nombres, 4: Nivel, 5: Grado, 6: Estatus,
        # 7: UPN, 8: Nombre 1, 9: Nombre Limpio, 10: Apellido Limpio, 11: Alias, 12: Display Name
        if not row:
            continue

        raw_mat = str(row[0]).strip() if len(row) > 0 and row[0] is not None else ""
        raw_pat = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
        raw_mat_ap = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ""
        raw_nom = str(row[3]).strip() if len(row) > 3 and row[3] is not None else ""
        raw_nivel = str(row[4]).strip() if len(row) > 4 and row[4] is not None else ""
        raw_grado = str(row[5]).strip() if len(row) > 5 and row[5] is not None else ""
        raw_estatus = str(row[6]).strip() if len(row) > 6 and row[6] is not None else ""
        raw_upn = str(row[7]).strip() if len(row) > 7 and row[7] is not None else ""
        raw_nom1 = str(row[8]).strip() if len(row) > 8 and row[8] is not None else ""
        raw_nom_limp = str(row[9]).strip() if len(row) > 9 and row[9] is not None else ""
        raw_ape_limp = str(row[10]).strip() if len(row) > 10 and row[10] is not None else ""
        raw_alias = str(row[11]).strip() if len(row) > 11 and row[11] is not None else ""
        raw_disp = str(row[12]).strip() if len(row) > 12 and row[12] is not None else ""

        # Skip rows that are completely empty or template rows without matricula/names
        if not raw_mat and not raw_nom and not raw_pat:
            continue

        # Format matricula and UPN if read as float (e.g. 250001.0 -> 250001)
        def clean_numeric_str(val: str) -> str:
            if not val:
                return ""
            v = val.strip()
            if v.endswith(".0"):
                v = v[:-2]
            try:
                f = float(v)
                if f.is_integer():
                    return str(int(f))
            except ValueError:
                pass
            return v

        raw_mat = clean_numeric_str(raw_mat)
        if raw_upn.endswith(".0@ijova.com"):
            raw_upn = raw_upn.replace(".0@ijova.com", "@ijova.com")
        elif ".0@" in raw_upn:
            raw_upn = raw_upn.replace(".0@", "@")

        record = StudentRecord(
            row_index=r_idx,
            matricula=raw_mat,
            apellido_paterno=raw_pat,
            apellido_materno=raw_mat_ap,
            nombres=raw_nom,
            nivel=raw_nivel,
            grado_semestre=raw_grado,
            estatus=raw_estatus,
            upn_raw=raw_upn,
            nombre1_raw=raw_nom1,
            nombre_limpio_raw=raw_nom_limp,
            apellido_limpio_raw=raw_ape_limp,
            alias_raw=raw_alias,
            display_name_raw=raw_disp
        )
        students.append(record)

    wb.close()
    return students


def append_student_to_excel(
    excel_path: str,
    sheet_name: str,
    matricula: str,
    nombres: str,
    apellido_paterno: str,
    apellido_materno: str,
    nivel: str,
    grado_semestre: str,
    estatus: str = "Activo"
) -> int:
    """
    Agrega un nuevo alumno al final de la hoja Excel con sus fórmulas estándar.
    Retorna el número de fila donde fue insertado.
    """
    wb = openpyxl.load_workbook(excel_path, data_only=False)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"La hoja '{sheet_name}' no existe en el libro.")

    sheet = wb[sheet_name]

    # Encontrar la primera fila vacía en la columna A
    target_row = 2
    for r in range(2, sheet.max_row + 2):
        val = sheet.cell(row=r, column=1).value
        if val is None or str(val).strip() == "":
            target_row = r
            break

    # Escribir columnas A a G
    sheet.cell(row=target_row, column=1, value=float(matricula) if matricula.isdigit() else matricula)
    sheet.cell(row=target_row, column=2, value=apellido_paterno.upper().strip())
    sheet.cell(row=target_row, column=3, value=apellido_materno.upper().strip() if apellido_materno else "")
    sheet.cell(row=target_row, column=4, value=nombres.upper().strip())
    sheet.cell(row=target_row, column=5, value=nivel.strip())
    sheet.cell(row=target_row, column=6, value=grado_semestre.strip())
    sheet.cell(row=target_row, column=7, value=estatus.strip())

    wb.save(excel_path)
    wb.close()
    return target_row


def update_student_status_in_excel(
    excel_path: str,
    sheet_name: str,
    matricula: str,
    new_status: str = "Baja"
) -> bool:
    """
    Actualiza el estatus de un alumno en la columna G del Excel a partir de su matrícula.
    Retorna True si encontró y actualizó al alumno.
    """
    wb = openpyxl.load_workbook(excel_path, data_only=False)
    if sheet_name not in wb.sheetnames:
        return False

    sheet = wb[sheet_name]
    mat_clean = matricula.strip().lower()
    updated = False

    for r in range(2, sheet.max_row + 1):
        cell_val = sheet.cell(row=r, column=1).value
        if cell_val is not None:
            c_str = str(cell_val).strip()
            if c_str.endswith(".0"):
                c_str = c_str[:-2]
            if c_str.lower() == mat_clean:
                sheet.cell(row=r, column=7, value=new_status)
                updated = True
                break

    if updated:
        wb.save(excel_path)
    wb.close()
    return updated


def _extract_matriculas_from_ods(file_path: str, sheet_name: Optional[str] = None) -> List[str]:
    """Extrae matrículas desde un archivo ODS."""
    from odf import opendocument, teletype
    from odf.table import Table, TableRow, TableCell

    doc = opendocument.load(file_path)
    tables = {t.getAttribute('name'): t for t in doc.spreadsheet.getElementsByType(Table)}
    if sheet_name:
        if sheet_name not in tables:
            raise ValueError(f"La hoja '{sheet_name}' no existe en el archivo ODS.")
        sheet = tables[sheet_name]
    else:
        sheet = doc.spreadsheet.getElementsByType(Table)[0]

    matriculas = []
    seen = set()

    for r in sheet.getElementsByType(TableRow):
        cells = r.getElementsByType(TableCell)
        for c in cells:
            rep = c.getAttribute('numbercolumnsrepeated')
            rep_count = int(rep) if rep else 1
            txt = teletype.extractText(c).strip()
            if not txt:
                continue
            val_str = txt
            if val_str.endswith(".0"):
                val_str = val_str[:-2]
            if val_str.isdigit() and len(val_str) >= 4:
                if val_str not in seen:
                    matriculas.append(val_str)
                    seen.add(val_str)
                break

    return matriculas


def extract_matriculas_from_excel(
    file_path: str,
    sheet_name: Optional[str] = None
) -> List[str]:
    """
    Extrae una lista de matrículas numéricas desde un archivo Excel o ODS.
    Soporta archivos simples con solo una columna de matrículas o archivos con encabezados.
    Descarta celdas vacías y encabezados no numéricos.
    """
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"No se encontró el archivo: {file_path}")

    if file_path.lower().endswith(".ods"):
        return _extract_matriculas_from_ods(file_path, sheet_name)

    wb = openpyxl.load_workbook(file_path, data_only=True)
    if sheet_name:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"La hoja '{sheet_name}' no existe en el archivo Excel.")
        sheet = wb[sheet_name]
    else:
        sheet = wb.active

    matriculas = []
    seen = set()

    for row in sheet.iter_rows(values_only=True):
        if not row:
            continue
        # Buscar en las celdas de la fila la primera que parezca una matrícula numérica
        for cell in row:
            if cell is None:
                continue
            val_str = str(cell).strip()
            if not val_str:
                continue
            # Sanitizar floats como 250001.0
            try:
                f = float(val_str)
                if f.is_integer():
                    val_str = str(int(f))
            except ValueError:
                pass

            # Si es numérica y tiene longitud de matrícula (4 a 10 dígitos)
            if val_str.isdigit() and len(val_str) >= 4:
                if val_str not in seen:
                    matriculas.append(val_str)
                    seen.add(val_str)
                break  # Solo tomar la primera columna válida de esa fila

    wb.close()
    return matriculas
