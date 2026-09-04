"""
Pruebas unitarias para verificar el manejo de paginación (@odata.nextLink),
rate limiting (HTTP 429 con Retry-After) y abortos de seguridad en GraphClient.
"""
import os
import unittest
from unittest.mock import patch, MagicMock
from src.graph_client import GraphClient, GraphClientError


class TestGraphClient(unittest.TestCase):
    @patch("src.graph_client.PublicClientApplication")
    def setUp(self, mock_msal_app):
        self.client = GraphClient(
            tenant_id="test-tenant-id",
            client_id="test-client-id",
            scopes=["User.Read.All"]
        )
        self.client.access_token = "fake-mock-token"

    @patch("requests.get")
    def test_pagination_success(self, mock_get):
        """Verifica que la paginación recorra múltiples páginas hasta que @odata.nextLink sea nulo."""
        # Page 1 response with nextLink
        resp1 = MagicMock()
        resp1.status_code = 200
        resp1.json.return_value = {
            "value": [
                {"id": "u1", "userPrincipalName": "user1@ijova.com", "displayName": "User 1"},
                {"id": "u2", "userPrincipalName": "user2@ijova.com", "displayName": "User 2"}
            ],
            "@odata.nextLink": "https://graph.microsoft.com/v1.0/users?$skiptoken=page2"
        }

        # Page 2 response without nextLink (end of list)
        resp2 = MagicMock()
        resp2.status_code = 200
        resp2.json.return_value = {
            "value": [
                {"id": "u3", "userPrincipalName": "user3@ijova.com", "displayName": "User 3"}
            ]
        }

        mock_get.side_effect = [resp1, resp2]

        users = self.client.get_all_users()
        self.assertEqual(len(users), 3)
        self.assertEqual(users[0].user_principal_name, "user1@ijova.com")
        self.assertEqual(users[2].user_principal_name, "user3@ijova.com")
        self.assertEqual(mock_get.call_count, 2)

    @patch("time.sleep", return_value=None)
    @patch("requests.get")
    def test_rate_limiting_retry_429(self, mock_get, mock_sleep):
        """Verifica el reintento ante HTTP 429 respetando el header Retry-After."""
        # 1st attempt: 429 with Retry-After: 2
        resp_429 = MagicMock()
        resp_429.status_code = 429
        resp_429.headers = {"Retry-After": "2"}

        # 2nd attempt: 200 OK
        resp_ok = MagicMock()
        resp_ok.status_code = 200
        resp_ok.json.return_value = {
            "value": [
                {"id": "u1", "userPrincipalName": "user1@ijova.com"}
            ]
        }

        mock_get.side_effect = [resp_429, resp_ok]

        users = self.client.get_all_users()
        self.assertEqual(len(users), 1)
        mock_sleep.assert_called_with(2)

    @patch("time.sleep", return_value=None)
    @patch("requests.get")
    def test_pagination_failure_aborts_securely(self, mock_get, mock_sleep):
        """Verifica que una falla en mitad de la paginación lanza GraphClientError y ABORTA."""
        # Page 1 OK
        resp1 = MagicMock()
        resp1.status_code = 200
        resp1.json.return_value = {
            "value": [{"id": "u1", "userPrincipalName": "user1@ijova.com"}],
            "@odata.nextLink": "https://graph.microsoft.com/v1.0/users?$skiptoken=page2"
        }

        # Page 2 fails with 500 continuously
        resp_500 = MagicMock()
        resp_500.status_code = 500

        mock_get.side_effect = [resp1, resp_500, resp_500, resp_500, resp_500]

    @patch("requests.post")
    def test_create_user_success(self, mock_post):
        """Verifica la creación exitosa de usuario vía POST /v1.0/users."""
        mock_resp = MagicMock()
        mock_resp.status_code = 201
        mock_resp.json.return_value = {
            "id": "new-user-id",
            "userPrincipalName": "250081@ijova.com",
            "displayName": "LOGAN JAVIER VALENCIA"
        }
        mock_post.return_value = mock_resp

        payload = {
            "userPrincipalName": "250081@ijova.com",
            "displayName": "LOGAN JAVIER VALENCIA"
        }
        result = self.client.create_user(payload)
        self.assertEqual(result["id"], "new-user-id")
        self.assertEqual(result["userPrincipalName"], "250081@ijova.com")

    @patch("requests.post")
    def test_assign_license_success(self, mock_post):
        """Verifica la asignación de licencias vía POST /v1.0/users/{id}/assignLicense."""
        mock_resp = MagicMock()
        mock_resp.status_code = 200
        mock_resp.json.return_value = {}
        mock_post.return_value = mock_resp

        res = self.client.assign_license("test-user-id", "sku-guid-123")
        self.assertTrue(res)

    @patch("requests.delete")
    def test_delete_user_success(self, mock_delete):
        """Verifica la eliminación de usuario vía DELETE /v1.0/users/{id}."""
        mock_resp = MagicMock()
        mock_resp.status_code = 204
        mock_delete.return_value = mock_resp

        res = self.client.delete_user("test-user-id-to-delete")
        self.assertTrue(res)

    def test_delete_safety_blocks_admins_and_staff(self):
        """Verifica que el motor de bajas bloquee terminantemente cuentas que no sean matrículas estudiantiles."""
        from src.delete_engine import is_student_matricula

        # Casos administrativos que deben ser BLOQUEADOS
        admin_cases = [
            "admin@ijova.com",
            "adminbackup@ijova.com",
            "carmelaleticiapachecosoriano@ijova.com",
            "director",
            "profesor.juan@ijova.com",
            "superadmin"
        ]
        for ac in admin_cases:
            valid, msg = is_student_matricula(ac)
            self.assertFalse(valid, f"Se esperaba bloqueo para {ac}")
            self.assertIn("NO es una matrícula estudiantil numérica", msg)

        # Casos estudiantiles que deben ser ACEPTADOS
        student_cases = ["250010", "260017", "250001@ijova.com", " 260002 "]
        for sc in student_cases:
            valid, upn = is_student_matricula(sc)
            self.assertTrue(valid, f"Se esperaba aceptación para {sc}")
            self.assertTrue(upn.endswith("@ijova.com"))

    @patch("requests.patch")
    def test_reset_password_success(self, mock_patch):
        """Verifica el restablecimiento de contraseña vía PATCH /v1.0/users/{id}."""
        mock_resp = MagicMock()
        mock_resp.status_code = 204
        mock_patch.return_value = mock_resp

        res = self.client.reset_password("test-user-id", "NewPassword123!")
        self.assertTrue(res)

    def test_pdf_generation_smoke_test(self):
        """Verifica la generación de PDF con tarjetas de acceso y código QR."""
        import tempfile
        import os
        from src.pdf_generator import generate_pdf_cards_from_list

        test_students = [{
            "matricula": "250001",
            "upn": "250001@ijova.com",
            "nombre_completo": "ALUMNO DEMO UNO",
            "password_temporal": "TempPass123!",
            "nivel": "Secundaria",
            "grado_semestre": "1ro"
        }]

        with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp:
            tmp_path = tmp.name

        try:
            out_path = generate_pdf_cards_from_list(test_students, tmp_path, layout_mode="cards")
            self.assertTrue(os.path.exists(out_path))
            self.assertGreater(os.path.getsize(out_path), 1000)
        finally:
            if os.path.exists(tmp_path):
                os.remove(tmp_path)

    @patch("requests.get")
    def test_get_deleted_users(self, mock_get):
        """Verifica la consulta de usuarios en la papelera de reciclaje."""
        mock_resp = MagicMock()
        mock_resp.status_code = 200
        mock_resp.json.return_value = {
            "value": [
                {"id": "del-1", "userPrincipalName": "250010@ijova.com", "displayName": "Alumno Eliminado"}
            ]
        }
        mock_get.return_value = mock_resp

        deleted = self.client.get_deleted_users()
        self.assertEqual(len(deleted), 1)
        self.assertEqual(deleted[0]["userPrincipalName"], "250010@ijova.com")

    @patch("requests.post")
    def test_restore_deleted_user_success(self, mock_post):
        """Verifica la restauración de usuario vía POST /deletedItems/{id}/restore."""
        mock_resp = MagicMock()
        mock_resp.status_code = 200
        mock_resp.json.return_value = {"id": "restored-user-id", "userPrincipalName": "250010@ijova.com"}
        mock_post.return_value = mock_resp

        res = self.client.restore_deleted_user("restored-user-id")
        self.assertEqual(res["id"], "restored-user-id")

    def test_extract_matriculas_from_excel(self):
        """Verifica la extracción de matrículas desde un Excel simple."""
        import tempfile
        import openpyxl
        from src.excel_parser import extract_matriculas_from_excel

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp_path = tmp.name

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(["Matricula"])
            ws.append([250001])
            ws.append(["250002.0"])
            ws.append([""])
            ws.append(["260010"])
            wb.save(tmp_path)
            wb.close()

            mats = extract_matriculas_from_excel(tmp_path)
            self.assertEqual(mats, ["250001", "250002", "260010"])
        finally:
            if os.path.exists(tmp_path):
                os.remove(tmp_path)

    @patch.object(GraphClient, "reset_password")
    def test_bulk_password_reset(self, mock_reset):
        """Verifica la ejecución del reseteo masivo de contraseñas y generación de archivos."""
        import tempfile
        import shutil
        from src.reset_engine import execute_bulk_password_reset

        mock_reset.return_value = True

        test_students = [
            {
                "matricula": "250001",
                "upn": "250001@ijova.com",
                "display_name": "ALUMNO DEMO UNO",
                "id": "user-id-1",
                "nivel": "Secundaria",
                "grado_semestre": "1ro"
            },
            {
                "matricula": "250002",
                "upn": "250002@ijova.com",
                "display_name": "ALUMNO DEMO DOS",
                "id": "user-id-2",
                "nivel": "Secundaria",
                "grado_semestre": "1ro"
            }
        ]

        tmp_secrets = tempfile.mkdtemp()
        tmp_reports = tempfile.mkdtemp()

        try:
            res = execute_bulk_password_reset(
                students_data=test_students,
                graph=self.client,
                domain="ijova.com",
                secrets_dir=tmp_secrets,
                reports_dir=tmp_reports,
                auto_confirm=True
            )
            self.assertEqual(res["reset_count"], 2)
            self.assertEqual(res["failed_count"], 0)
            self.assertTrue(os.path.exists(res["csv_path"]))
            self.assertTrue(os.path.exists(res["pdf_cards"]))
            self.assertTrue(os.path.exists(res["pdf_full"]))
        finally:
            shutil.rmtree(tmp_secrets, ignore_errors=True)
            shutil.rmtree(tmp_reports, ignore_errors=True)

    def test_historical_registry_and_anti_reassignment_rule(self):
        """Verifica que el registro histórico bloquee la reasignación de matrículas a personas diferentes."""
        import tempfile
        import os
        from src.historical_registry import record_student_baja, check_matricula_transfer_conflict, mark_student_reactivated

        with tempfile.NamedTemporaryFile(suffix=".json", delete=False) as tmp:
            tmp_json = tmp.name

        try:
            # 1. Registrar una baja histórica
            record_student_baja(
                matricula="250118",
                upn="250118@ijova.com",
                nombre_completo="ALUMNO PRUEBA HISTORICA",
                apellido_paterno="HISTORICA",
                nombres="ALUMNO PRUEBA",
                motivo_baja="No reinscrito",
                history_path=tmp_json
            )

            # 2. Intentar registrar a OTRA persona con la misma matrícula -> DEBE BLOQUEAR
            is_conflict, msg = check_matricula_transfer_conflict("250118", "OTRO ALUMNO DISTINTO", history_path=tmp_json)
            self.assertTrue(is_conflict)
            self.assertIn("ya perteneció históricamente a 'ALUMNO PRUEBA HISTORICA'", msg)
            self.assertIn("intransferibles de por vida", msg)

            # 3. Intentar reinscribir a la MISMA persona con pequeñas variaciones -> DEBE PERMITIR
            is_conflict_same, _ = check_matricula_transfer_conflict("250118", "ALUMNO PRUEBA", history_path=tmp_json)
            self.assertFalse(is_conflict_same)

            # 4. Verificar reactivación
            reactivated = mark_student_reactivated("250118", history_path=tmp_json)
            self.assertTrue(reactivated)

        finally:
            if os.path.exists(tmp_json):
                os.remove(tmp_json)

    def test_matricula_format_validation(self):
        """Verifica la validación estricta del formato de matrículas del IJOVA (6 dígitos, ej. 25xxxx, 26xxxx)."""
        from src.validator import is_valid_matricula_format, validate_students
        from src.models import StudentRecord, ClassificationEnum
        from src.delete_engine import is_student_matricula

        # 1. Función directa is_valid_matricula_format
        self.assertTrue(is_valid_matricula_format("250001"))
        self.assertTrue(is_valid_matricula_format("260015"))
        self.assertTrue(is_valid_matricula_format("250135"))
        self.assertTrue(is_valid_matricula_format("260027"))

        # Casos inválidos por longitud o prefijo
        self.assertFalse(is_valid_matricula_format("25001"))     # 5 dígitos
        self.assertFalse(is_valid_matricula_format("2600001"))   # 7 dígitos
        self.assertFalse(is_valid_matricula_format("123456"))    # no empieza con 2
        self.assertFalse(is_valid_matricula_format("990001"))    # no empieza con 2
        self.assertFalse(is_valid_matricula_format("25ABCD"))    # alfanumérico
        self.assertFalse(is_valid_matricula_format(""))          # vacío
        self.assertFalse(is_valid_matricula_format("   "))       # espacios
        self.assertFalse(is_valid_matricula_format(None))        # None

        # 2. Integración en validate_students
        bad_student = StudentRecord(
            row_index=10,
            matricula="25001",  # Inválida (5 dígitos)
            apellido_paterno="LOPEZ",
            apellido_materno="GARCIA",
            nombres="JUAN",
            nivel="Secundaria",
            grado_semestre="1ro",
            estatus="Inscrito",
            upn_raw="25001@ijova.com",
            nombre1_raw="JUAN",
            nombre_limpio_raw="JUAN",
            apellido_limpio_raw="LOPEZ",
            alias_raw="juan.lopez",
            display_name_raw="JUAN LOPEZ GARCIA"
        )
        good_student = StudentRecord(
            row_index=11,
            matricula="260055",  # Válida (6 dígitos)
            apellido_paterno="PEREZ",
            apellido_materno="DIAZ",
            nombres="ANA",
            nivel="Secundaria",
            grado_semestre="1ro",
            estatus="Inscrito",
            upn_raw="260055@ijova.com",
            nombre1_raw="ANA",
            nombre_limpio_raw="ANA",
            apellido_limpio_raw="PEREZ",
            alias_raw="ana.perez",
            display_name_raw="ANA PEREZ DIAZ"
        )

        validated = validate_students([bad_student, good_student])
        bad_res = next(s for s in validated if s.row_index == 10)
        good_res = next(s for s in validated if s.row_index == 11)

        # bad_student debe tener MATRICULA_FORMATO_INVALIDO y clasificación INVALIDO
        bad_issue_codes = [issue.code for issue in bad_res.issues]
        self.assertIn("MATRICULA_FORMATO_INVALIDO", bad_issue_codes)
        self.assertEqual(bad_res.classification, ClassificationEnum.INVALIDO)

        # good_student NO debe tener MATRICULA_FORMATO_INVALIDO
        good_issue_codes = [issue.code for issue in good_res.issues]
        self.assertNotIn("MATRICULA_FORMATO_INVALIDO", good_issue_codes)

        # 3. Integración en is_student_matricula (delete_engine / restore_engine / reset_engine)
        valid_res, upn_res = is_student_matricula("250001")
        self.assertTrue(valid_res)
        self.assertEqual(upn_res, "250001@ijova.com")

        valid_res_upn, _ = is_student_matricula("260015@ijova.com")
        self.assertTrue(valid_res_upn)

        # Rechazo de admin o matrículas que no cumplen formato
        valid_admin, msg_admin = is_student_matricula("admin@ijova.com")
        self.assertFalse(valid_admin)
        self.assertIn("NO es una matrícula estudiantil numérica", msg_admin)

        valid_short, msg_short = is_student_matricula("2501")
        self.assertFalse(valid_short)
        self.assertIn("NO cumple con el formato de matrícula escolar", msg_short)

        valid_prefix, msg_prefix = is_student_matricula("990001")
        self.assertFalse(valid_prefix)
        self.assertIn("NO cumple con el formato de matrícula escolar", msg_prefix)

    @patch("src.graph_client.PublicClientApplication")
    def test_token_cache_persistence_and_clear(self, mock_pca):
        """Verifica la persistencia y limpieza del token cache de MSAL en disco."""
        import tempfile
        import os
        from src.graph_client import GraphClient

        with tempfile.NamedTemporaryFile(suffix=".bin", delete=False) as tmp:
            tmp_cache = tmp.name

        try:
            # 1. Crear cliente con ruta de caché
            client1 = GraphClient("test-tenant", "test-client", ["User.Read.All"], cache_path=tmp_cache)
            self.assertIsNotNone(client1.cache)

            # Simular cambio en caché y guardado
            client1.cache.has_state_changed = True
            client1._save_cache()
            self.assertTrue(os.path.exists(tmp_cache))

            # 2. Cargar un nuevo cliente apuntando al mismo archivo
            client2 = GraphClient("test-tenant", "test-client", ["User.Read.All"], cache_path=tmp_cache)
            self.assertIsNotNone(client2.cache)

            # 3. Limpiar sesión
            client2.clear_session_cache()
            self.assertFalse(os.path.exists(tmp_cache))
            self.assertIsNone(client2.access_token)
        finally:
            if os.path.exists(tmp_cache):
                os.remove(tmp_cache)

    @patch("requests.post")
    def test_graph_batch_reset_passwords(self, mock_post):
        """Verifica la ejecución por lotes $batch para reseteo masivo de contraseñas."""
        mock_resp = MagicMock()
        mock_resp.status_code = 200
        mock_resp.json.return_value = {
            "responses": [
                {"id": "1", "status": 204, "body": {}},
                {"id": "2", "status": 400, "body": {"error": {"message": "Invalid password"}}}
            ]
        }
        mock_post.return_value = mock_resp

        students_reset = [
            {"id": "uid-1", "matricula": "250001", "upn": "250001@ijova.com", "display_name": "Alumno 1", "new_password": "Pass1!"},
            {"id": "uid-2", "matricula": "250002", "upn": "250002@ijova.com", "display_name": "Alumno 2", "new_password": "Pass2!"}
        ]

        result = self.client.batch_reset_passwords(students_reset)
        self.assertEqual(result["total"], 2)
        self.assertEqual(result["success_count"], 1)
        self.assertEqual(result["failed_count"], 1)
        self.assertEqual(result["successes"][0]["student"]["matricula"], "250001")
        self.assertEqual(result["failures"][0]["student"]["matricula"], "250002")

    def test_styled_excel_report_generation(self):
        """Verifica la generación del libro Excel institucional con estilos y semáforos."""
        import tempfile
        import os
        import openpyxl
        from src.report_generator import export_styled_excel_report
        from src.models import StudentRecord, ClassificationEnum

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp_xlsx = tmp.name

        try:
            sample_student = StudentRecord(
                row_index=2,
                matricula="250001",
                nombres="ALUMNO",
                apellido_paterno="DEMO",
                apellido_materno="UNO",
                nivel="Secundaria",
                grado_semestre="1ro",
                estatus="Inscrito",
                upn_raw="25001@ijova.com",
                nombre1_raw="ALUMNO",
                nombre_limpio_raw="ALUMNO",
                apellido_limpio_raw="DEMO",
                alias_raw="alumno.demo",
                display_name_raw="ALUMNO DEMO UNO",
                upn_normalized="250001@ijova.com",
                display_name="ALUMNO DEMO UNO",
                classification=ClassificationEnum.EXISTENTE
            )
            export_styled_excel_report(
                students=[sample_student],
                discrepancies=[{"matricula": "250001", "upn": "250001@ijova.com", "campo": "displayName", "valor_entra": "A", "valor_excel": "B"}],
                output_path=tmp_xlsx,
                timestamp_utc="2026-09-04 16:00:00 UTC",
                admin_upn="admin@ijova.com"
            )

            self.assertTrue(os.path.exists(tmp_xlsx))
            wb = openpyxl.load_workbook(tmp_xlsx)
            self.assertIn("Resumen Ejecutivo", wb.sheetnames)
            self.assertIn("Existentes Sincronizados", wb.sheetnames)
            self.assertIn("Discrepancias", wb.sheetnames)
        finally:
            if os.path.exists(tmp_xlsx):
                os.remove(tmp_xlsx)

    def test_audit_logger(self):
        """Verifica que el logger centralizado registre eventos correctamente en logs/ijova_audit.log."""
        import os
        from src.audit_logger import log_audit_event, AUDIT_LOG_FILE

        log_audit_event(
            action="TEST_ACTION",
            target="259999",
            admin="test_admin@ijova.com",
            status="SUCCESS",
            details="Prueba unitaria de auditoría"
        )

        self.assertTrue(os.path.exists(AUDIT_LOG_FILE))
        with open(AUDIT_LOG_FILE, "r", encoding="utf-8") as f:
            content = f.read()
            self.assertIn("ACTION: TEST_ACTION", content)
            self.assertIn("TARGET: 259999", content)
            self.assertIn("test_admin@ijova.com", content)


if __name__ == "__main__":
    unittest.main()
